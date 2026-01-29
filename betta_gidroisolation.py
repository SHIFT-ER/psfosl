
import openpyxl
import os
from openpyxl.drawing.image import Image  
from openpyxl.styles import Font, PatternFill, Border, Alignment
from datetime import datetime
import openpyxl.styles
import openpyxl.utils
from openpyxl.worksheet.pagebreak import Break
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
import numpy as np
import tempfile



class Gidroisolation_report():


    def __init__(self):
        self.result_file_name = "Отчёт_по_гидроизоляции.xlsx"
        self.page_name = "Участок"
        self.object_name = "Объект"#
        self.client_name = "Заказчик"#
        self.contract_name = "Договор"#
        self.device_name = "Прибор"
        self.zav_number = "777"
        self.device_valide_until = "01.01.01"
        self.plot_location = "Распололжение контролируемого участка"
        self.work_date = "01.01.01"
        self.plot_name = "Название слоя"
        self.square = 25
        self.values = [1,2,3,4]
        self.normative_value = 0.1
        self.current_date = datetime.today().strftime("%d.%m.%y") # Получаем текущую дату
        self.sootv_status = "Соответствет"
        self.velichina_adgezi = []
        self.last_cell = ""

    def __init__(self, r_f = "Отчёт_по_гидроизоляции.xlsx", p_n = "Участок",
                 o_n = "Объект", client_n = "Заказчик", contract_n = "Договор",
                 d_n = "Прибор",  z_n = "777", d_v_u = "01.01.01",
                 p_l = "Распололжение контролируемого участка", w_d = "01.01.01", 
                 p_name = "Название слоя", s = 25, v = [1,2,3,4]):
        self.result_file_name = r_f
        self.page_name = p_n
        self.object_name = o_n
        self.client_name = client_n
        self.contract_name = contract_n
        self.device_name = d_n
        self.zav_number = z_n
        self.device_valide_until = d_v_u
        self.plot_location = p_l
        self.work_date = w_d
        self.plot_name = p_name
        self.square = s
        self.values = v
        self.normative_value = 0.1 
        self.current_date = datetime.today().strftime("%d.%m.%y") # Получаем текущую дату
        self.sootv_status = "соответствет"
        self.velichina_adgezi = []
        self.last_cell = ""
        self.table_end_row = 0  # Последняя строка таблицы с результатами
        # Скрипт по сохранению файла
        # result_ws.protection.sheet = False  # Отключаем защиту листа
        # result_wb.security.lockStructure = False  # Отключаем защиту книги
        # desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")  
        # file_path = os.path.join(desktop_path, result_file_name)
        # result_wb.save(file_path)

    def set_filename(self, f):
        self.set_filename = f

    # =(значение_силы*1000)/(площадь*100)
    def count_adgezi(self, val:float):
        x = (val*1000)/(self.square*100)
        self.velichina_adgezi.append(x)
        return x
    
    # =СРЗНАЧ(M67:N70)
    def count_average_adgezi(self):
        y = []
        for i in self.values: y.append(self.count_adgezi(i))
        x = round((sum(y)/len(self.values)), 2)
        x = str(x).replace(".", ",")
        return x
    
    def get_average_adgezi_value(self):
        """Возвращает среднее значение адгезии как число (не строку)"""
        # Вычисляем величины адгезии если они ещё не вычислены
        if len(self.velichina_adgezi) == 0:
            for val in self.values:
                self.count_adgezi(val)
        
        if len(self.velichina_adgezi) == 0:
            return 0.0
        
        return sum(self.velichina_adgezi) / len(self.velichina_adgezi)
    
    def create_adhesion_chart(self):
        """Создаёт график 'Величина адгезии vs. № участка' и возвращает путь к файлу изображения"""
        # Убеждаемся, что величины адгезии вычислены
        # (они должны быть вычислены в set_tables)
        # Используем уже вычисленные значения, чтобы не дублировать
        if len(self.velichina_adgezi) == 0:
            # Если по какой-то причине значения не вычислены, вычисляем их
            for val in self.values:
                self.count_adgezi(val)
        
        if len(self.velichina_adgezi) == 0:
            raise ValueError("Нет данных для построения графика")
        
        adhesion_values = self.velichina_adgezi.copy()
        
        # Номера участков (X)
        plot_numbers = list(range(1, len(adhesion_values) + 1))
        
        # Среднее значение
        average_value = self.get_average_adgezi_value()
        
        # Нормативное значение
        normative_value = float(self.normative_value) if self.normative_value else 0.1
        
        # ±15% диапазон
        minus_15_percent = average_value * 0.85
        plus_15_percent = average_value * 1.15
        
        # Создание графика
        plt.figure(figsize=(10, 6))
        
        # Основной график - величина адгезии
        plt.plot(plot_numbers, adhesion_values, 'o-', linewidth=2.5, markersize=10, 
                label='Величина адгезии', color='#0066CC', markerfacecolor='#0066CC', 
                markeredgecolor='white', markeredgewidth=1.5)
        
        # Нормативное значение (горизонтальная линия)
        plt.axhline(y=normative_value, color='red', linestyle='--', linewidth=2.5, 
                   label=f'Нормативное значение ({normative_value} МПа)', alpha=0.8)
        
        # Среднее значение (горизонтальная линия)
        plt.axhline(y=average_value, color='green', linestyle='--', linewidth=2.5, 
                   label=f'Среднее значение ({average_value:.2f} МПа)', alpha=0.8)
        
        # ±15% диапазон (заштрихованная область)
        plt.axhspan(minus_15_percent, plus_15_percent, alpha=0.25, color='orange', 
                   label=f'±15% диапазон ({minus_15_percent:.2f} - {plus_15_percent:.2f} МПа)')
        
        # Линии границ ±15%
        plt.axhline(y=minus_15_percent, color='orange', linestyle=':', linewidth=2, alpha=0.8)
        plt.axhline(y=plus_15_percent, color='orange', linestyle=':', linewidth=2, alpha=0.8)
        
        # Настройка осей
        plt.xlabel('№ участка', fontsize=13, fontweight='bold')
        plt.ylabel('Величина адгезии, МПа', fontsize=13, fontweight='bold')
        plt.title('Величина адгезии vs. № участка', fontsize=15, fontweight='bold', pad=15)
        plt.grid(True, alpha=0.4, linestyle='--', linewidth=0.8)
        plt.legend(loc='upper right', fontsize=10, framealpha=0.9, shadow=True)
        
        # Установка меток на оси X как целые числа
        plt.xticks(plot_numbers, fontsize=11)
        plt.yticks(fontsize=11)
        
        # Улучшение внешнего вида
        plt.tight_layout(pad=2.0)
        
        # Устанавливаем стиль для более профессионального вида
        plt.style.use('default')
        # Добавляем рамку вокруг графика
        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(1.2)
        ax.spines['bottom'].set_linewidth(1.2)
        
        # Сохранение графика во временный файл
        temp_dir = tempfile.gettempdir()
        chart_path = os.path.join(temp_dir, f'adhesion_chart_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        return chart_path

    def create_empty_report(self, page_name = "Отчёт", organisation_header_image_path = None, formula_image_path = None, template_path = None):
        # Ищем файл шаблона
        if template_path and os.path.exists(template_path):
            top_page_name = template_path
        else:
            # Ищем файл шаблона в текущей директории
            script_dir = os.path.dirname(os.path.abspath(__file__))
            top_page_name = os.path.join(script_dir, "gidroisolation_top.xlsx")
        
        # Загружаем исходный файл
        if not os.path.exists(top_page_name):
            # Если шаблона нет, создаём пустой файл
            result_wb = openpyxl.Workbook()
            result_ws = result_wb.active
            result_ws.title = page_name
        else:
            source_wb = openpyxl.load_workbook(top_page_name)
            source_ws = source_wb.active  # Получаем лист с изображением

            # Создаём новый файл
            result_wb = openpyxl.Workbook()
            result_ws = result_wb.active
            result_ws.title = page_name

            # Копируем данные (текст, числа и формулы)
            # Копируем данные и стили (включая курсив)
            result_ws.protection.sheet = False

            for row in source_ws.iter_rows():
                for cell in row:
                    # Получаем ячейку в новом файле, создавая её автоматически
                    new_cell = result_ws.cell(row=cell.row, column=cell.column)  
                    new_cell.value = cell.value  # Копируем значение

                    # Копируем шрифт (включая курсив)
                    if cell.font:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )

                    # Копируем заполнение (если есть)
                    if cell.fill:
                        new_cell.fill = PatternFill(
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                            fill_type=cell.fill.fill_type
                        )

                    # Копируем границы (если есть)
                    if cell.border:
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                            diagonal=cell.border.diagonal,
                            diagonal_direction=cell.border.diagonal_direction,
                            outline=cell.border.outline,
                            vertical=cell.border.vertical,
                            horizontal=cell.border.horizontal
                        )

                    # Копируем выравнивание (если есть)
                    if cell.alignment:
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )

            # Копируем ширину вскех столбцов
            for col_letter, col_dimension in source_ws.column_dimensions.items():
                result_ws.column_dimensions[col_letter].width = col_dimension.width

            #Копируем обхединения ячеек
            for merged_range in source_ws.merged_cells.ranges:
                result_ws.merge_cells(str(merged_range))
        
        # Вставляем изображения, если пути переданы
        if organisation_header_image_path and os.path.exists(organisation_header_image_path):
            try:
                organisation_header_image = Image(organisation_header_image_path)  
                organisation_header_image.width = 800
                organisation_header_image.height = 175
                result_ws.add_image(organisation_header_image, "C1")
            except Exception as e:
                print(f"Ошибка при вставке шапки организации: {e}")

        if formula_image_path and os.path.exists(formula_image_path):
            try:
                gidroisolation_formula_image = Image(formula_image_path)
                gidroisolation_formula_image.width = 145
                gidroisolation_formula_image.height = 70
                result_ws.add_image(gidroisolation_formula_image, "K50")
            except Exception as e:
                print(f"Ошибка при вставке формулы: {e}")


        
        result_ws.protection.sheet = False  # Отключаем защиту листа
        result_wb.security.lockStructure = False  # Отключаем защиту книги
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")  
        file_path = os.path.join(desktop_path, self.result_file_name)
        result_wb.save(file_path)

    def create_conclusion(self):
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        current_filename = os.path.join(desktop_path, self.result_file_name)
        result_wb = openpyxl.load_workbook(current_filename)  # Загружаем существующий файл
        result_ws = result_wb.active

        # Вычисляем позицию выводов после графика
        # График начинается на строке self.table_end_row + 2 (заголовок на строке table_end_row + 1, график на table_end_row + 2)
        # График занимает примерно 23 строки (высота графика 456 пикселей / ~20 пикселей на строку)
        # Добавляем отступ после графика
        if hasattr(self, 'table_end_row') and self.table_end_row > 0:
            chart_start_row = self.table_end_row + 2  # Заголовок на table_end_row + 1, график на table_end_row + 2
            chart_end_row = chart_start_row + 23  # График занимает ~23 строки
            row = chart_end_row + 2  # Отступ после графика
        else:
            # Если по какой-то причине table_end_row не установлен, используем старую логику
            row = 67 + len(self.values) + 3
        
        cell = "C"+str(row)
        result_ws[cell] = "4.ВЫВОДЫ."
        result_ws[cell].font = Font(bold=True, size=14)
        result_ws.row_breaks.append(Break(id=row))


        row += 1
        cell = "C"+str(row)
        start_cell = "C"+str(row)
        end_cell = "R"+str(row+1)
        r = start_cell + ":" + end_cell
        result_ws.merge_cells(r) 
        result_ws.row_dimensions[row].height = 50
        result_ws[cell] = '="4.1. В рамках выполнения работ по договору "&U29&" с "&U27&" на объекте по адресу: "&U25&" проведены работы по испытанию сцепления гидроизоляционных материалов."'
        result_ws[cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        result_ws[cell].font = Font(size = 14)

        row += 2
        cell = "C"+str(row)
        start_cell = "C"+str(row)
        end_cell = "R"+str(row+1)
        r = start_cell + ":" + end_cell
        result_ws.merge_cells(r) 
        result_ws.row_dimensions[row].height = 50
        controlir_uchast = "контролируемом участке"#если их потом будет несколько, то нужно будет поменять на "контролируемых участках"
        result_ws[cell] = f'4.2. По результатам проведённых работ, приведенных в таблице 1 установлено: что фактическое значение велечины адгезии испытанного материала на {controlir_uchast} составляет от {min(self.velichina_adgezi)} до {max(self.velichina_adgezi)} МПа, что {self.sootv_status} требованиям СП 71.13330.2017 и номративной документации."'
        result_ws[cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        result_ws[cell].font = Font(size = 14)

        row += 3
        cell = "D" + str(row)
        result_ws.merge_cells(f'D{row}:H{row}')
        result_ws[cell] = f'=D37'
        result_ws[cell].font = Font(size=14)
        self.last_cell = f'R{row}'


        result_ws.protection.sheet = False  # Отключаем защиту листа
        result_wb.security.lockStructure = False  # Отключаем защиту книги
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")  
        file_path = os.path.join(desktop_path, self.result_file_name)
        result_wb.save(file_path)

    
    def set_tables(self):
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        current_filename = os.path.join(desktop_path, self.result_file_name)
    ################################ Заполняем постоянные данные
        result_wb = openpyxl.load_workbook(current_filename)  # Загружаем существующий файл
        result_ws = result_wb.active
        ####Заполняем шапку обхекта

        result_ws["U25"] = self.object_name
        result_ws["U27"] = self.client_name
        result_ws["U29"] = self.contract_name
        
        ####Заполняем шапку прибора
        result_ws["U35"] = self.device_name
        result_ws["V35"] = self.zav_number
        result_ws["W35"] = self.device_valide_until
        result_ws["M22"] = self.current_date

        # Инициализация столбцов и начальной строки
        pillars = 'CDEFGHIJKLMNOPQRSTUVW'
        start_row = 67

        # Заполнение столбца "номер участка испытаний"
        for i in range(len(self.values)):
            result_ws.row_dimensions[start_row + i].height = 30  # Увеличиваем высоту строк
            explore_place_number_cell = pillars[0] + str(start_row + i)
            result_ws[explore_place_number_cell] = i + 1
            result_ws[explore_place_number_cell].alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение столбца "Расположение контролируемого участка"
        cell_start = pillars[1] + str(start_row)
        cell_end = pillars[2] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)  # Объединяем ячейки
        result_ws[cell_start] = self.plot_location  # Записываем значение только в первую ячейку объединённой области
        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)

        # Заполнение столбца "Дата проведения работ"
        cell_start = pillars[3] + str(start_row)
        cell_end = pillars[3] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)  # Объединение ячеек
        result_ws[cell_start] = self.work_date  # Записываем значение только в первую ячейку
        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)

        # Заполнение столбца "Испытываемый слой гидроизоляции"
        cell_start = pillars[4] + str(start_row)
        cell_end = pillars[5] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)  # Объединение ячеек
        result_ws[cell_start] = self.plot_name  # Записываем значение только в первую ячейку
        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)


        # Заполнение столбца "Площадь отрыва, см2"
        cell_start = pillars[6] + str(start_row)
        cell_end = pillars[7] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)  # Объединяем ячейки
        result_ws[cell_start] = self.square  # Записываем значение только в первую ячейку объединённой области
        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)

        # Заполнение столбца "Значение силы, кН"
        for i in range(len(self.values)):
            cell_start = pillars[8] + str(start_row+i)
            cell_end = pillars[9] + str(start_row + i)
            cells_range = cell_start + ":" + cell_end
            result_ws.merge_cells(cells_range)
            cell = pillars[8] + str(start_row + i)
            result_ws[cell] = self.values[i]
            result_ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение столбца "Величина адгезии, МПа"
        for i in range(len(self.values)):
            cell_start = pillars[10] + str(start_row + i)
            cell_end = pillars[11] + str(start_row + i)
            cells_range = cell_start + ":" + cell_end
            result_ws.merge_cells(cells_range)
            cell_equal_strength_value = pillars[8] + str(start_row + i)
            cell_square_value = "$" + pillars[6] + "$" + str(start_row)
            # print("=(" +  cell_equal_strength_value + "*1000)/(" + cell_square_value + "*100)")
            # =(K67*1000)/($I$67*100)
            result_ws[cell_start] = self.count_adgezi(self.values[i])
            result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение столбца "Среднее значение адгезии, МПа"
        # cell_start_values = pillars[10] + str(start_row)
        # cell_end_values = pillars[11] + str(start_row + len(self.values) - 1)
        cell_start = pillars[12] + str(start_row)
        cell_end = pillars[13] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)

        result_ws[cell_start]= str(self.count_average_adgezi())


        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)

        # Заполнение столбца "Нормативное значение, МПа"
        cell_start = pillars[14] + str(start_row)
        cell_end = pillars[14] + str(start_row + len(self.values) - 1)
        cells_range = cell_start + ":" + cell_end
        result_ws.merge_cells(cells_range)
        result_ws[cell_start] = self.normative_value
        result_ws[cell_start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_ws[cell_start].font = Font(size=12)
        
        # Заполнение столбца "Соответствие нормативной документации"
        for i in range(len(self.values)):
            cell = pillars[15] + str(start_row + i)
            x = float(self.values[i])/(float(self.square)*100)
            if (x < float(self.normative_value)): result_ws[cell] = "Соответствует"
            else: 
                result_ws[cell] ="Не соответствует"
                self.sootv_status = "не соответствует"
            # cell = pillars[15] + str(start_row + i)
            # result_ws[cell] = '=_xludf.ЕСЛИ(M67 < 0,1: "Соответствует";"Не соответствует")'
            result_ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение столбца "-15% от средней адгезии"
        for i in range(len(self.values)):
            cell = pillars[16] + str(start_row + i)
            result_ws[cell] = '=$' + pillars[12] + "$" + str(start_row) + "-($" + pillars[12] + "$" + str(start_row) + "*15%)"
            result_ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение столбца "+15% от средней адгезии"
        for i in range(len(self.values)):
            cell = pillars[17] + str(start_row + i)
            result_ws[cell] = '=$' + pillars[12] + "$" + str(start_row) + "+($" + pillars[12] + "$" + str(start_row) + "*15%)"
            result_ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        
        result_ws.row_dimensions[65].height = 55
        
        # Сохраняем информацию о последней строке таблицы для вставки графика
        self.table_end_row = start_row + len(self.values) - 1

        result_ws.protection.sheet = False  # Отключаем защиту листа
        result_wb.security.lockStructure = False  # Отключаем защиту книги
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")  
        file_path = os.path.join(desktop_path, self.result_file_name)
        result_wb.save(file_path)
    
    def insert_chart_into_report(self):
        """Вставляет график адгезии в отчёт после таблицы с результатами"""
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        current_filename = os.path.join(desktop_path, self.result_file_name)
        result_wb = openpyxl.load_workbook(current_filename)
        result_ws = result_wb.active
        
        # Создаём график
        chart_path = None
        try:
            chart_path = self.create_adhesion_chart()
            
            # Вычисляем позицию для вставки графика динамически
            # Таблица заканчивается на строке self.table_end_row
            # Вставляем график через 2 строки после таблицы
            chart_start_row = self.table_end_row + 2
            
            # Определяем столбец для вставки (столбец C, как в таблице)
            chart_column = 'C'
            chart_cell = f'{chart_column}{chart_start_row}'
            
            # Вставляем график
            if chart_path and os.path.exists(chart_path):
                chart_image = Image(chart_path)
                # Устанавливаем размер графика (в пикселях)
                # Размер подбирается так, чтобы график хорошо смотрелся в Excel
                chart_image.width = 760  # Немного меньше 800 для лучшего отображения
                chart_image.height = 456  # Сохраняем соотношение 10:6
                result_ws.add_image(chart_image, chart_cell)
                
                # Добавляем заголовок графика перед графиком
                title_cell = f'{chart_column}{chart_start_row - 1}'
                result_ws[title_cell] = "3. График зависимости величины адгезии от номера участка"
                result_ws[title_cell].font = Font(bold=True, size=12)
                
                # Увеличиваем высоту строки для заголовка
                result_ws.row_dimensions[chart_start_row - 1].height = 25
                
                # Увеличиваем высоту строк для графика (примерно на высоту графика в единицах Excel)
                # Высота графика ~456 пикселей, высота строки Excel ~20 пикселей по умолчанию
                # Нужно примерно 23 строки для графика
                for i in range(23):
                    row_num = chart_start_row + i
                    if row_num not in result_ws.row_dimensions or result_ws.row_dimensions[row_num].height is None:
                        result_ws.row_dimensions[row_num].height = 20
                
        except Exception as e:
            print(f"Ошибка при создании и вставке графика: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # Сохраняем файл перед удалением временного файла графика
            result_ws.protection.sheet = False
            result_wb.security.lockStructure = False
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_path = os.path.join(desktop_path, self.result_file_name)
            result_wb.save(file_path)
            
            # Удаляем временный файл графика после сохранения
            if chart_path and os.path.exists(chart_path):
                try:
                    os.remove(chart_path)
                except Exception as e:
                    print(f"Не удалось удалить временный файл графика: {e}")

    def create_print_area(self):
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        current_filename = os.path.join(desktop_path, self.result_file_name)
        result_wb = openpyxl.load_workbook(current_filename)  
        result_ws = result_wb.active

        result_ws.print_area = f"C3:R41,C42:{self.last_cell}"
        result_ws.page_setup.orientation = result_ws.ORIENTATION_LANDSCAPE

        result_ws.page_margins.left = 0.2
        result_ws.page_margins.right = 0.2
        # result_ws.page_margins.top = 0.2
        # result_ws.page_margins.bottom = 0.2
        # result_ws.page_margins.header = 0.1
        # result_ws.page_margins.footer = 0.1

        # result_ws.col_breaks.append(Break(id=17))
        

        result_ws.protection.sheet = False  # Отключаем защиту листа
        result_wb.security.lockStructure = False  # Отключаем защиту книги
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")  
        file_path = os.path.join(desktop_path, self.result_file_name)
        result_wb.save(file_path)





    def create_gidroisolation_report(self, name = "Отчёт", organisation_header_image_path = None, formula_image_path = None, template_path = None):
        self.create_empty_report(name, organisation_header_image_path, formula_image_path, template_path)
        self.set_tables()
        self.insert_chart_into_report()  # Вставляем график после таблицы
        self.create_conclusion()
        self.create_print_area()













